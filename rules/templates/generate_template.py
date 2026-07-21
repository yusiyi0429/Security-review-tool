#!/usr/bin/env python3
"""Generate the Security Review Rules Excel template."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "security-review-rules-template.xlsx"

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, columns):
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_data(ws, row_idx, col_count):
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_sheet(wb, name, columns, sample_rows):
    ws = wb.active if wb.active and wb.active.title == "Sheet" and len(wb.sheetnames) == 1 else wb.create_sheet()
    ws.title = name
    style_header(ws, columns)

    for row_idx, row_data in enumerate(sample_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        style_data(ws, row_idx, len(columns))

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    return ws


def main():
    wb = openpyxl.Workbook()

    # ---- Categories ----
    add_sheet(wb, "Categories",
        columns=[
            ("CategoryId", 14),
            ("Name", 30),
            ("Description", 50),
            ("Enabled", 10),
        ],
        sample_rows=[
            ["SENS-001", "Sensitive Personal Data", "Scan for PII, credentials, and tokens", True],
            ["SENS-002", "Access Control", "Detect improper ACLs and permissions", True],
        ],
    )

    # ---- Assets ----
    add_sheet(wb, "Assets",
        columns=[
            ("AssetTypeId", 14),
            ("Name", 30),
            ("Description", 50),
            ("FocusWeights", 50),
        ],
        sample_rows=[
            [
                "ASSET-001",
                "Source Code Repository",
                "Git repositories containing application source code",
                '{"SENS-001": 1.0, "SENS-002": 0.5}',
            ],
        ],
    )

    # ---- ComplianceRules ----
    add_sheet(wb, "ComplianceRules",
        columns=[
            ("Id", 20),
            ("AssetTypeId", 14),
            ("Name", 30),
            ("Description", 50),
            ("EvidenceField", 25),
            ("RequiredStatus", 25),
        ],
        sample_rows=[
            [
                "CR-SOURCE-001",
                "ASSET-001",
                "GitHub Branch Protection",
                "Verify branch protection rules are enabled for the default branch",
                "branch_protection",
                "enabled",
            ],
        ],
    )

    # ---- Rules ----
    add_sheet(wb, "Rules",
        columns=[
            ("RuleId", 18),
            ("CategoryId", 14),
            ("FindingKind", 20),
            ("Severity", 12),
            ("DetectionConfidence", 20),
            ("DetectorId", 18),
            ("DetectorConfigId", 25),
            ("AppliesToAssets", 40),
            ("RequiresSemanticReview", 22),
            ("Enabled", 10),
        ],
        sample_rows=[
            [
                "RULE-AWS-KEY-001",
                "SENS-001",
                "SensitiveContent",
                "Critical",
                "High",
                "DET-CRED-001",
                "aws-access-key",
                "ASSET-001,ASSET-002",
                False,
                True,
            ],
        ],
    )

    # ---- Detectors ----
    add_sheet(wb, "Detectors",
        columns=[
            ("DetectorId", 18),
            ("Kind", 22),
            ("ConfigId", 25),
            ("Parameters", 60),
            ("MaxMatchesPerChunk", 20),
        ],
        sample_rows=[
            [
                "DET-CRED-001",
                "KnownFormat",
                "aws-access-key",
                '{"pattern": "AKIA[0-9A-Z]{16}", "description": "AWS Access Key ID"}',
                100,
            ],
        ],
    )

    # Remove default sheet if present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(OUTPUT)
    print(f"Template written to {OUTPUT}")


if __name__ == "__main__":
    main()
